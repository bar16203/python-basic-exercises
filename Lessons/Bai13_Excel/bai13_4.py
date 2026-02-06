from openpyxl import load_workbook, Workbook
wb = load_workbook("products.xlsx")
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(["Name", "Price", "Stock", "Total"])

for row in ws.iter_rows(min_row = 2, values_only = True): 
    name, price, qty = row
    total = price*qty
    new_ws.append([name, price, qty, total])
new_wb.save("result.xlsx")