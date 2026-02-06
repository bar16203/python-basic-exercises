from openpyxl import load_workbook, Workbook

wb = load_workbook("result.xlsx")
ws = wb.active
sum = 0

for row in ws.iter_rows(min_row=2, min_col= 4, max_col= 4, values_only= True): 
    sum += row[0]
print("Sum: ", sum)